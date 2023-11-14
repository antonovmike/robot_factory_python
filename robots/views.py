# import logging

from django.views import View
from django.http import FileResponse
from django.db.models import Count
from django.utils import timezone
from openpyxl import Workbook

from rest_framework import generics
from rest_framework.response import Response

from .models import Robot
from .serializers import RobotSerializer


class RobotCreateView(generics.CreateAPIView):
    queryset = Robot.objects.all()
    serializer_class = RobotSerializer


class ReportGenerator:
    def __init__(self):
        self.week_ago = timezone.now() - timezone.timedelta(weeks=1)
        self.robots = Robot.objects.filter(created__gte=self.week_ago)
        self.data = self.robots.values('model', 'version').annotate(
            total=Count('model')).order_by('model', 'version')

    def generate_report(self):
        return self.data


class WorkbookCreator:
    def __init__(self, data):
        self.data = data
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    def create_workbook(self):
        if not self.data:
            ws = self.wb.create_sheet(title="No data")
            ws['A1'] = "No data for the last week"
        else:
            current_letter = ''
            for row_data in self.data:
                letter = row_data['model'][0]
                if letter != current_letter:
                    current_letter = letter
                    ws = self.wb.create_sheet(title=letter)
                    headers = ["Модель", "Версия", "Количество за неделю"]
                    for col_num, column_title in enumerate(headers, 1):
                        col_letter = ws.cell(row=1, column=col_num).column_letter
                        ws['{}1'.format(col_letter)] = column_title
                        ws.column_dimensions[col_letter].width = 15
                row_num = ws.max_row + 1
                ws.cell(row=row_num, column=1, value=row_data['model'])
                ws.cell(row=row_num, column=2, value=row_data['version'])
                ws.cell(row=row_num, column=3, value=row_data['total'])
        return self.wb


class FileHandler:
    def __init__(self, wb):
        self.wb = wb

    def save_workbook_to_file(self):
        self.wb.save('robots_report.xlsx')
        return 'robots_report.xlsx'

    def create_file_response(self, filepath):
        response = FileResponse(
            open(filepath, 'rb'),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=robots_report.xlsx'
        return response


class RobotReportView(View):
    def get(self, request):
        report_generator = ReportGenerator()
        data = report_generator.generate_report()

        # FIX IT
        total_robots = Robot.objects.all().count()
        print(f"Total amount of robots on {timezone.now().date()} is {total_robots}")
        # logging.info(f"Total amount of robots on {timezone.now().date()} is {total_robots}")

        workbook_creator = WorkbookCreator(data)
        wb = workbook_creator.create_workbook()

        file_handler = FileHandler(wb)
        filepath = file_handler.save_workbook_to_file()
        response = file_handler.create_file_response(filepath)

        return response


class RobotDeleteView(generics.DestroyAPIView):
    queryset = Robot.objects.all()
    serializer_class = RobotSerializer
    lookup_field = 'serial'

    def delete(self, request, *args, **kwargs):
        # Get a robot by serial
        robot = self.get_object()

        model = robot.model
        version = robot.version
        serial = robot.serial

        robot.delete()

        return Response({"message": f"Робот {model} {version} с серийным номером {serial} удален."})
